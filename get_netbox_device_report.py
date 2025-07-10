#!/usr/bin/env python3

import os
import sys
import requests
import json
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles.numbers import BUILTIN_FORMATS
    import pytz
    from datetime import datetime
except ImportError:
    print("Required libraries not installed. Run 'pip install openpyxl pytz' and retry.", file=sys.stderr)
    sys.exit(1)

NETBOX_URL = os.environ.get('NETBOX_URL') or os.environ.get('NETBOX_API')
NETBOX_TOKEN = os.environ.get('NETBOX_TOKEN')
EXCLUDED_ROLE_IDS = {2, 11}

if not NETBOX_URL or not NETBOX_TOKEN:
    print("Missing NETBOX_URL (or NETBOX_API) or NETBOX_TOKEN environment variables.", file=sys.stderr)
    sys.exit(1)

headers = {
    'Authorization': f"Token {NETBOX_TOKEN}",
    'Accept': 'application/json',
}

HEADING_ORDER = [
    "Gateway/Router",
    "Switches",
    "Physical Hosts",
    "Virtual Machines",
    "Storage Area Network",
    "Network Attached Storage",
    "Production Workstations",
    "Uninterruptible Power Supply",
    "Printers",
    "Wireless Access Equipment",
]

DEVICE_ROLE_GROUPS = {
    "Gateway/Router": {
        "Enterprise": [12],
        "Operational": [34],
    },
    "Switches": {
        "Enterprise Core": [5],
        "Enterprise Edge": [1],
        "Operational": [28],
    },
    "Physical Hosts": {
        "Enterprise": [6],
        "Operational": [43],
    },
    "Virtual Machines": {
        "Enterprise": [4, 19, 17, 20, 18, 33],
        "Operational": [35, 36, 37, 38, 39, 40],
    },
    "Storage Area Network": {
        "Enterprise": [16],
    },
    "Network Attached Storage": {
        "Enterprise": [15],
        "Operational": [41],
    },
    "Production Workstations": {
        "Operational": [30],
        "Quality Assurance": [32],
        "Optimisation": [31],
    },
    "Uninterruptible Power Supply": {
        "Enterprise": [14],
        "Operational": [44],
    },
    "Printers": {
        "Enterprise": [24],
        "Operational": [26],
    },
    "Wireless Access Equipment": {
        "Access Points": [10],
        "Point to Point": [45],
        "Access Equipment": [46],
    },
}

def get_heading_and_subheading(role_id):
    for heading, sub_map in DEVICE_ROLE_GROUPS.items():
        for subheading, role_ids in sub_map.items():
            if role_id in role_ids:
                return heading, subheading
    return "Other", "Other"

def tick(val):
    if val:
        return "✓", "00AA00"  # green
    else:
        return "✗", "FF0000"  # red

def short_desc(desc, length=100):
    if desc and len(desc) > length:
        return desc[:length-3] + "..."
    return desc or ""

site_device_counts = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {'count': 0, 'devices': []})))
device_debug_file = "/runner/device_debug.json"
with open(device_debug_file, "w") as dbg:
    dbg.write("")

def fetch_netbox_items(url, item_type):
    all_items = []
    while url:
        r = requests.get(url, headers=headers, timeout=30)
        r.raise_for_status()
        result = r.json()
        for item in result.get('results', []):
            item['_item_type'] = item_type
            all_items.append(item)
        url = result.get('next')
    return all_items

devices_url = f"{NETBOX_URL.rstrip('/')}/api/dcim/devices/?limit=1000&expand=role,site,tenant,contact,location,platform"
devices = fetch_netbox_items(devices_url, "device")

vms_url = f"{NETBOX_URL.rstrip('/')}/api/virtualization/virtual-machines/?limit=1000&expand=role,site,tenant,contact,location,platform"
vms = fetch_netbox_items(vms_url, "vm")

all_items = devices + vms

for item in all_items:
    with open(device_debug_file, "a") as dbg:
        dbg.write(json.dumps(item, indent=2) + "\n\n")
    site = item.get('site', {}).get('name', 'Unassigned Site')
    status = item.get('status', {}).get('value') if isinstance(item.get('status'), dict) else item.get('status')
    if status not in ('active', 1):
        continue
    role_id = None
    item_role = item.get('role', None)
    if isinstance(item_role, dict):
        role_id = item_role.get('id', None)
    elif isinstance(item_role, int):
        role_id = item_role
    else:
        role_id = None
    if role_id in EXCLUDED_ROLE_IDS:
        continue
    cf = item.get('custom_fields', {}) or {}

    # Compliance checks
    tenant_present = bool(item.get('tenant'))
    contact_present = bool(item.get('contact'))
    location_present = bool(item.get('location'))
    platform_present = bool(item.get('platform'))

    device_info = {
        'site': site,
        'heading': None,
        'subheading': None,
        'name': item.get('name', 'Unknown Device'),
        'description': item.get('description', ''),
        'tenant': item.get('tenant', {}).get('name') if isinstance(item.get('tenant'), dict) else item.get('tenant'),
        'contact': item.get('contact', {}).get('name') if isinstance(item.get('contact'), dict) else item.get('contact'),
        'location': item.get('location', {}).get('name') if isinstance(item.get('location'), dict) else item.get('location'),
        'platform': item.get('platform', {}).get('name') if isinstance(item.get('platform'), dict) else item.get('platform'),
        'primary_ip': item.get('primary_ip'),
        'serial': item.get('serial'),
        'backup_primary': cf.get('last_backup_data_prim'),
        'monitoring_required': cf.get('mon_required'),
        # Extra checks
        'tenant_present': tenant_present,
        'contact_present': contact_present,
        'location_present': location_present,
        'platform_present': platform_present,
    }
    if role_id is not None:
        heading, subheading = get_heading_and_subheading(role_id)
        device_info['heading'] = heading
        device_info['subheading'] = subheading
        site_device_counts[site][heading][subheading]['count'] += 1
        site_device_counts[site][heading][subheading]['devices'].append(device_info)
    else:
        device_info['heading'] = "Other"
        device_info['subheading'] = "Other"
        site_device_counts[site]['Other']['Other']['count'] += 1
        site_device_counts[site]['Other']['Other']['devices'].append(device_info)

# ---- Excel Generation: One sheet per site, table per heading+subheading ----
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove default sheet

headers = [
    "Device Name", "Description", "Tenant", "Contact", "Location", "Platform",
    "Primary IP", "Serial", "Backup Data - Primay", "Monitoring Required"
]

header_fill = PatternFill("solid", fgColor="00336699")
header_font = Font(bold=True, color="FFFFFFFF")
title_font = Font(bold=True, size=14)
subhead_font = Font(bold=True, color="333399")

# ---- Add Summary Worksheet ----
summary_ws = wb.create_sheet(title="Summary", index=0)
summary_headers = [
    "Site", "Heading", "Subheading", "Device Count",
    "% Tenant ✓", "% Contact ✓", "% Location ✓", "% Platform ✓",
    "% Primary IP ✓", "% Serial ✓", "% Backup ✓", "% Monitoring ✓"
]
summary_ws.append(summary_headers)
for cell in summary_ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

summary_rows = []
for site in sorted(site_device_counts):
    for heading in HEADING_ORDER + ["Other"]:
        if heading not in site_device_counts[site]:
            continue
        sub_map = DEVICE_ROLE_GROUPS.get(heading, {"Other": []}) if heading != "Other" else {"Other": []}
        for subheading in (sub_map if heading != "Other" else ["Other"]):
            devices = site_device_counts[site][heading][subheading]['devices'] if subheading in site_device_counts[site][heading] else []
            count = len(devices)
            if count == 0:
                continue
            tenant_tick = sum(1 for d in devices if d.get('tenant_present'))
            contact_tick = sum(1 for d in devices if d.get('contact_present'))
            location_tick = sum(1 for d in devices if d.get('location_present'))
            platform_tick = sum(1 for d in devices if d.get('platform_present'))
            ip_tick = sum(1 for d in devices if d.get('primary_ip'))
            serial_tick = sum(1 for d in devices if d.get('serial'))
            backup_tick = sum(1 for d in devices if d.get('backup_primary'))
            monitor_tick = sum(1 for d in devices if d.get('monitoring_required') is not False)
            row = [
                site,
                heading,
                subheading,
                count,
                (tenant_tick/count) if count else 0,
                (contact_tick/count) if count else 0,
                (location_tick/count) if count else 0,
                (platform_tick/count) if count else 0,
                (ip_tick/count) if count else 0,
                (serial_tick/count) if count else 0,
                (backup_tick/count) if count else 0,
                (monitor_tick/count) if count else 0,
            ]
            summary_ws.append(row)
            summary_rows.append(row)

# ---- Add Totals Across Sites Row ----
total_count = 0
total_tenant_tick = 0
total_contact_tick = 0
total_location_tick = 0
total_platform_tick = 0
total_ip_tick = 0
total_serial_tick = 0
total_backup_tick = 0
total_monitor_tick = 0

for row in summary_rows:
    count = int(row[3] or 0)
    total_count += count
    try:
        total_tenant_tick   += count * float(row[4] or 0)
        total_contact_tick  += count * float(row[5] or 0)
        total_location_tick += count * float(row[6] or 0)
        total_platform_tick += count * float(row[7] or 0)
        total_ip_tick       += count * float(row[8] or 0)
        total_serial_tick   += count * float(row[9] or 0)
        total_backup_tick   += count * float(row[10] or 0)
        total_monitor_tick  += count * float(row[11] or 0)
    except Exception:
        pass

if total_count > 0:
    total_row = [
        "ALL SITES",
        "",
        "",
        total_count,
        total_tenant_tick/total_count,
        total_contact_tick/total_count,
        total_location_tick/total_count,
        total_platform_tick/total_count,
        total_ip_tick/total_count,
        total_serial_tick/total_count,
        total_backup_tick/total_count,
        total_monitor_tick/total_count
    ]
else:
    total_row = ["ALL SITES", "", "", 0, 0, 0, 0, 0, 0, 0, 0, 0]

summary_ws.append(total_row)
for col_idx in range(1, summary_ws.max_column + 1):
    cell = summary_ws.cell(row=summary_ws.max_row, column=col_idx)
    cell.font = Font(bold=True, color="00336699")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# ---- Set Percentage Format ----
for row in summary_ws.iter_rows(min_row=2, max_row=summary_ws.max_row, min_col=5, max_col=12):
    for cell in row:
        cell.number_format = '0.0%'

# ---- Conditional Formatting for Compliance % Columns ----
last_row = summary_ws.max_row
for col_letter in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
    summary_ws.conditional_formatting.add(
        f"{col_letter}2:{col_letter}{last_row}",
        CellIsRule(operator='lessThan', formula=['0.8'], fill=PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid'))
    )
    summary_ws.conditional_formatting.add(
        f"{col_letter}2:{col_letter}{last_row}",
        CellIsRule(operator='between', formula=['0.8', '0.95'], fill=PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'))
    )
    summary_ws.conditional_formatting.add(
        f"{col_letter}2:{col_letter}{last_row}",
        CellIsRule(operator='greaterThanOrEqual', formula=['0.95'], fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'))
    )

# ---- Add Report Generation Date/Time (AEST/AEDT) ----
tz = pytz.timezone("Australia/Melbourne")
dt_row = summary_ws.max_row + 2
report_dt = datetime.now(tz).strftime('%Y-%m-%d %H:%M:%S %Z')
summary_ws.cell(row=dt_row, column=1, value=f"Report generated: {report_dt}")
summary_ws.merge_cells(start_row=dt_row, start_column=1, end_row=dt_row, end_column=summary_ws.max_column)
summary_ws[f"A{dt_row}"].font = Font(italic=True, color="888888")

# Auto-size columns for summary
for col_idx in range(1, summary_ws.max_column + 1):
    col_letter = get_column_letter(col_idx)
    max_length = 0
    for row in summary_ws.iter_rows(min_row=1, max_row=summary_ws.max_row, min_col=col_idx, max_col=col_idx):
        for cell in row:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
    summary_ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

# ---- Per-Site Worksheets ----
for site in sorted(site_device_counts):
    ws = wb.create_sheet(title=site[:31])
    ws.append([f"{site} Device Report"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws['A1'].font = title_font
    rownum = 2

    for heading in HEADING_ORDER:
        if heading not in site_device_counts[site]:
            continue
        for subheading in DEVICE_ROLE_GROUPS[heading]:
            devices = site_device_counts[site][heading][subheading]['devices']
            if not devices:
                continue
            ws.append([f"{heading} - {subheading}"])
            ws.merge_cells(start_row=rownum, start_column=1, end_row=rownum, end_column=len(headers))
            ws[f'A{rownum}'].font = subhead_font
            rownum += 1

            # Table header
            for colidx, head in enumerate(headers, 1):
                cell = ws.cell(row=rownum, column=colidx, value=head)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            rownum += 1

            # Devices for this role/subheading
            for device in sorted(devices, key=lambda d: d['name']):
                ws.append([
                    device['name'],
                    short_desc(device.get('description', ''), 100),
                    device.get('tenant', ''),
                    device.get('contact', ''),
                    device.get('location', ''),
                    device.get('platform', ''),
                    "", "", "", "", ""  # Placeholders for ticks/crosses
                ])
                current_row = ws.max_row
                # Ticks: Tenant, Contact, Location, Platform, Primary IP, Serial, Backup, Monitoring
                tick_tenant = tick(device.get('tenant_present'))
                tick_contact = tick(device.get('contact_present'))
                tick_location = tick(device.get('location_present'))
                tick_platform = tick(device.get('platform_present'))
                tick_primary_ip = tick(device.get('primary_ip'))
                tick_serial = tick(device.get('serial'))
                tick_backup = tick(device.get('backup_primary'))
                tick_monitor = tick(device.get('monitoring_required') is not False)
                for idx, (value, color) in enumerate(
                    [tick_tenant, tick_contact, tick_location, tick_platform,
                     tick_primary_ip, tick_serial, tick_backup, tick_monitor], start=7
                ):
                    cell = ws.cell(row=current_row, column=idx)
                    cell.value = value
                    cell.font = Font(color=color, bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                rownum += 1

            ws.append([])
            rownum += 1

    # ---- Add "Other - Other" table if present ----
    other_devices = site_device_counts[site].get('Other', {}).get('Other', {}).get('devices', [])
    if other_devices:
        ws.append(["Other - Other"])
        ws.merge_cells(start_row=rownum, start_column=1, end_row=rownum, end_column=len(headers))
        ws[f'A{rownum}'].font = subhead_font
        rownum += 1

        for colidx, head in enumerate(headers, 1):
            cell = ws.cell(row=rownum, column=colidx, value=head)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        rownum += 1

        for device in sorted(other_devices, key=lambda d: d['name']):
            ws.append([
                device['name'],
                short_desc(device.get('description', ''), 100),
                device.get('tenant', ''),
                device.get('contact', ''),
                device.get('location', ''),
                device.get('platform', ''),
                "", "", "", "", ""
            ])
            current_row = ws.max_row
            tick_tenant = tick(device.get('tenant_present'))
            tick_contact = tick(device.get('contact_present'))
            tick_location = tick(device.get('location_present'))
            tick_platform = tick(device.get('platform_present'))
            tick_primary_ip = tick(device.get('primary_ip'))
            tick_serial = tick(device.get('serial'))
            tick_backup = tick(device.get('backup_primary'))
            tick_monitor = tick(device.get('monitoring_required') is not False)
            for idx, (value, color) in enumerate(
                [tick_tenant, tick_contact, tick_location, tick_platform,
                 tick_primary_ip, tick_serial, tick_backup, tick_monitor], start=7
            ):
                cell = ws.cell(row=current_row, column=idx)
                cell.value = value
                cell.font = Font(color=color, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            rownum += 1
        ws.append([])
        rownum += 1

    # Auto-size columns (skip merged cells)
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

excel_file = "/runner/cmdb_device_report.xlsx"
wb.save(excel_file)
print("Excel report generated:", excel_file)
